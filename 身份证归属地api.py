# app_id:rgihdrm0kslojqvm
# https://www.mxnzp.com/api/idcard/search?idcard=your&app_secret=your&app_id=your
# app_secret:WnhrK251TWlUUThqaVFWbG5OeGQwdz09

# 获取表格身份证号数据
from time import sleep

from openpyxl import load_workbook
import json


wb = load_workbook(fr'././data/姓名身份证.xlsx')  # 获取已存在的工作簿
print("获取表格" + "成功")
ws = wb.active  # 获取第一张工作表对象
cols = ws.columns
prjTuple = tuple(cols)
id_list = []
for idcard in prjTuple[1][1:]:
    id_list.append(idcard.value)
print(id_list)
# api获取json
import requests as re
head = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
}
for idx in range(0, len(id_list)):
    url = f'https://www.mxnzp.com/api/idcard/search?idcard={id_list[idx]}&app_secret=WnhrK251TWlUUThqaVFWbG5OeGQwdz09&app_id=rgihdrm0kslojqvm'
    result = re.get(url=url, headers=head).json()
    sleep(1.1)
    print(result)
    # 取出相应信息

    # JSON数据
    json_data = json.dumps(result)

    # 将JSON数据解析为Python字典
    data = json.loads(json_data)

    # 提取各个字段的值
    code, msg, idcard_num, address, birthday, sex = \
        data["code"], data["msg"], data["data"]["idCardNum"], data["data"]["address"], \
        data["data"]["birthday"], data["data"]["sex"]

    if code != 1:
        sleep(2)
        continue


    # 插入到excel
    ws.cell(idx + 2, 3).value, ws.cell(idx + 2, 4).value, ws.cell(idx + 2, 5).value = birthday, sex, address

wb.save('./data/test.xlsx')



